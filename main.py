import os
import mysql.connector
import pandas as pd
from dotenv import load_dotenv
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

# .envファイルから環境変数を読み込む
load_dotenv()

# MySQLの接続情報を環境変数から取得
db_config = {
    'host': os.getenv('MYSQL_HOST'),
    'user': os.getenv('MYSQL_USER'),
    'password': os.getenv('MYSQL_PASSWORD'),
    'database': os.getenv('MYSQL_DATABASE'),
}

# MySQLに接続
conn = mysql.connector.connect(**db_config)
cursor = conn.cursor()

# テーブル一覧を取得
cursor.execute("SHOW TABLES;")
tables = cursor.fetchall()

# Excelファイルを作成
excel_writer = pd.ExcelWriter('db_design.xlsx', engine='openpyxl', mode='w')

# 各テーブルの情報をExcelに書き込む
for table in tables:
    table_name = table[0]
    
    # テーブルのカラム情報を取得
    cursor.execute(f"DESCRIBE {table_name};")
    columns = cursor.fetchall()
    
    # カラム情報をDataFrameに変換
    column_df = pd.DataFrame(columns, columns=['フィールド名', '型', 'Null許可', 'Key', 'Default', 'Extra'])
    
    # DataFrameをExcelに書き込み（3行目から）
    start_row = 2  # 3行目にデータを書き込む
    column_df.to_excel(excel_writer, sheet_name=table_name, index=False, startrow=start_row)
    
    worksheet = excel_writer.sheets[table_name]
    worksheet['A1'] = 'テーブル名'
    worksheet['B1'] = table_name

    # ヘッダーセルのスタイル設定
    header_style = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    font_style = Font(color="FFFFFF")

    for cell in worksheet[1]:
        cell.fill = header_style
        cell.font = font_style

    # カラムごとに列の幅を調整
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        try:
            max_length = max(len(str(cell.value)) for cell in column)
        except:
            pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Excelファイルを保存
# excel_writer.save()

# Close the ExcelWriter object
excel_writer.close()

# MySQLとの接続を閉じる
cursor.close()
conn.close()

